"""
文件职责：在内存解码账单文件，不保存原文件或执行表格公式。
主要内容：UTF-8/GB18030 CSV 解码、单工作表 XLSX 读取和资源上限。
关键边界：拒绝加密文件、宏、公式、多表及压缩炸弹；输出保留工作表行号的文本。
"""

import csv
import io
from datetime import date, datetime
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook

MAX_BYTES = 10 * 1024 * 1024


def decode_statement(name: str, raw: bytes) -> str:
    """只接受明示格式；不执行用户文件、不解密或请求支付密码。"""
    if not raw or len(raw) > MAX_BYTES:
        raise ValueError("文件为空或超过 10 MB")
    if name.lower().endswith(".csv"):
        for encoding in ("utf-8-sig", "gb18030"):
            try:
                content = raw.decode(encoding)
                if "\x00" in content:
                    raise ValueError("文件不是文本账单")
                if len(content.encode()) > MAX_BYTES:
                    raise ValueError("解码后的账单超过 10 MB")
                return content
            except UnicodeDecodeError:
                continue
        raise ValueError("无法识别文本编码，请保留原文件")
    if not name.lower().endswith(".xlsx"):
        raise ValueError("支持 CSV 或 XLSX；压缩包请在本机解密解压，不要提供支付密码")
    try:
        with ZipFile(io.BytesIO(raw)) as archive:
            entries = archive.infolist()
            if len(entries) > 200 or sum(e.file_size for e in entries) > 30 * 1024 * 1024:
                raise ValueError("工作簿解压大小超过限制")
            if any(e.flag_bits & 1 for e in entries):
                raise ValueError("不支持加密工作簿")
            if any("vbaproject" in e.filename.lower() for e in entries):
                raise ValueError("不支持宏工作簿")
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=False, keep_links=False)
        try:
            if len(workbook.worksheets) != 1:
                raise ValueError("请提供单工作表账单")
            sheet = workbook.worksheets[0]
            sheet.reset_dimensions()
            stream = io.StringIO()
            writer = csv.writer(stream)
            for index, cells in enumerate(sheet.iter_rows(), start=1):
                if index > 5200 or len(cells) > 80:
                    raise ValueError("工作表超过行列上限")
                values: list[str] = []
                for cell in cells:
                    if cell.data_type == "f":
                        raise ValueError("不支持公式单元格，请使用原始导出账单")
                    value = cell.value
                    if isinstance(value, datetime):
                        values.append(value.isoformat(sep=" ", timespec="seconds"))
                    elif isinstance(value, date):
                        values.append(value.isoformat())
                    elif isinstance(value, (int, float)) and abs(value) >= 10**15:
                        raise ValueError("长交易编号被保存为数值，可能已丢失精度")
                    else:
                        values.append("" if value is None else str(value))
                while values and not values[-1]:
                    values.pop()
                writer.writerow(values)
                if stream.tell() > MAX_BYTES:
                    raise ValueError("工作表文本超过限制")
            content = stream.getvalue()
            if len(content.encode()) > MAX_BYTES:
                raise ValueError("工作表文本超过限制")
            return content
        finally:
            workbook.close()
    except (BadZipFile, KeyError, OSError) as exc:
        raise ValueError("工作簿损坏或已加密，请保留原文件") from exc
